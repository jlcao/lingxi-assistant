#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
版本提测冒烟检查清单 Excel 生成器
生成 WPS 直接可用的美化版 Excel 表格
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# 定义颜色
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # 蓝色表头
TITLE_FILL = PatternFill(start_color="203864", end_color="203864", fill_type="solid")  # 深蓝色标题
LIGHT_BLUE_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")  # 浅蓝背景
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # 白色
WARNING_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 浅红警告

# 字体
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="微软雅黑", size=10)

# 对齐
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center", wrap_text=True)

# 边框
THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
)


def set_column_widths(ws, widths):
    """设置列宽"""
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def apply_table_style(ws, start_row, end_row, max_col):
    """应用表格样式"""
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if row > start_row:
                cell.fill = WHITE_FILL
                cell.alignment = CENTER_ALIGN
                cell.font = NORMAL_FONT


def create_version_sheet(ws):
    """创建表 1：版本单"""
    ws.title = "版本单"
    
    # 标题
    ws.merge_cells('A1:L1')
    title = ws.cell(row=1, column=1, value="版本提测冒烟检查清单 - 版本单")
    title.fill = TITLE_FILL
    title.font = TITLE_FONT
    title.alignment = CENTER_ALIGN
    
    # 表头
    headers = [
        "版本号", "分支名称", "开发负责人", "指定 QA", "预计提测时间",
        "功能/交易说明", "风险等级", "是否兼容旧版", "回滚方案", "备注"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # 设置列宽
    set_column_widths(ws, [12, 18, 12, 10, 15, 30, 12, 15, 20, 20])
    
    # 添加数据验证（下拉框）
    yes_no_dv = DataValidation(type="list", formula1='"是，否"', allow_blank=True)
    yes_no_dv.error = "请选择是或否"
    ws.add_data_validation(yes_no_dv)
    yes_no_dv.add(ws.cell(row=4, column=8))
    
    risk_dv = DataValidation(type="list", formula1='"高，中，低"', allow_blank=True)
    ws.add_data_validation(risk_dv)
    risk_dv.add(ws.cell(row=4, column=7))
    
    # 添加空白行供填写
    for row in range(4, 10):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            cell.font = NORMAL_FONT
            if col == 8:  # 是否兼容旧版列
                yes_no_dv.add(cell)
            if col == 7:  # 风险等级列
                risk_dv.add(cell)
    
    # 冻结首行
    ws.freeze_panes = 'A4'


def create_files_sheet(ws):
    """创建表 2：文件列表"""
    ws.title = "文件列表"
    
    # 标题
    ws.merge_cells('A1:H1')
    title = ws.cell(row=1, column=1, value="修改文件列表")
    title.fill = TITLE_FILL
    title.font = TITLE_FONT
    title.alignment = CENTER_ALIGN
    
    # 表头 - 修改文件
    headers1 = ["序号", "文件名称/路径", "文件类型", "修改内容", "影响交易", "影响功能", "是否公共文件", "备注"]
    
    for col, header in enumerate(headers1, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    set_column_widths(ws, [6, 35, 12, 25, 20, 20, 15, 20])
    
    # 数据验证
    yes_no_dv = DataValidation(type="list", formula1='"是，否"', allow_blank=True)
    ws.add_data_validation(yes_no_dv)
    
    # 添加空白行
    for row in range(4, 12):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            cell.font = NORMAL_FONT
            if col in [5, 6, 7]:  # 影响交易、影响功能、是否公共文件
                yes_no_dv.add(cell)
    
    # 新增文件标题
    ws.merge_cells('A14:H14')
    title2 = ws.cell(row=14, column=1, value="新增文件列表")
    title2.fill = TITLE_FILL
    title2.font = TITLE_FONT
    title2.alignment = CENTER_ALIGN
    
    # 表头 - 新增文件
    headers2 = ["序号", "文件名称/路径", "文件类型", "用途", "归属交易/功能", "备注"]
    
    for col, header in enumerate(headers2, 1):
        cell = ws.cell(row=16, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # 添加空白行
    for row in range(17, 25):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            cell.font = NORMAL_FONT
    
    ws.freeze_panes = 'A4'


def create_api_sheet(ws):
    """创建表 3：接口清单"""
    ws.title = "接口清单"
    
    # 标题
    ws.merge_cells('A1:I1')
    title = ws.cell(row=1, column=1, value="接口清单")
    title.fill = TITLE_FILL
    title.font = TITLE_FONT
    title.alignment = CENTER_ALIGN
    
    # 表头
    headers = ["序号", "接口名称", "接口地址", "请求方式", "是否改动", "改动内容", "单元测试", "覆盖率", "影响交易", "备注"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    set_column_widths(ws, [6, 20, 30, 12, 10, 25, 15, 12, 20, 20])
    
    # 数据验证
    yes_no_dv = DataValidation(type="list", formula1='"是，否"', allow_blank=True)
    ws.add_data_validation(yes_no_dv)
    
    method_dv = DataValidation(type="list", formula1='"GET,POST,PUT,DELETE,PATCH"', allow_blank=True)
    ws.add_data_validation(method_dv)
    
    # 添加空白行
    for row in range(4, 15):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            cell.font = NORMAL_FONT
            if col == 5:  # 是否改动
                yes_no_dv.add(cell)
            if col == 4:  # 请求方式
                method_dv.add(cell)
    
    ws.freeze_panes = 'A4'


def create_test_sheet(ws):
    """创建表 4：交易开发与测试截图清单"""
    ws.title = "交易测试"
    
    # 标题
    ws.merge_cells('A1:H1')
    title = ws.cell(row=1, column=1, value="交易开发与测试截图清单")
    title.fill = TITLE_FILL
    title.font = TITLE_FONT
    title.alignment = CENTER_ALIGN
    
    # 表头
    headers = ["序号", "交易名称", "所属分支", "代码提交截图", "后台发网关报文", "CD 报文", "Service 流水", "交易测试截图", "备注"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    set_column_widths(ws, [6, 20, 15, 15, 18, 15, 15, 18, 20])
    
    # 数据验证
    yes_no_dv = DataValidation(type="list", formula1='"有，无"', allow_blank=True)
    ws.add_data_validation(yes_no_dv)
    
    # 添加空白行
    for row in range(4, 12):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            cell.font = NORMAL_FONT
            if col in [4, 5, 6, 7, 8]:  # 截图相关列
                yes_no_dv.add(cell)
    
    ws.freeze_panes = 'A4'


def create_qa_sheet(ws):
    """创建表 5：公共组件检查和 QA 检查"""
    ws.title = "QA 检查"
    
    # 标题
    ws.merge_cells('A1:F1')
    title = ws.cell(row=1, column=1, value="公共组件/公共接口专项检查 & QA 检查结果")
    title.fill = TITLE_FILL
    title.font = TITLE_FONT
    title.alignment = CENTER_ALIGN
    
    # === 公共组件修改 ===
    ws.merge_cells('A3:F3')
    section1 = ws.cell(row=3, column=1, value="【公共组件修改检查】")
    section1.fill = LIGHT_BLUE_FILL
    section1.font = Font(name="微软雅黑", size=11, bold=True)
    section1.alignment = LEFT_ALIGN
    section1.border = THIN_BORDER
    
    headers1 = ["组件名称", "影响交易列表", "验证交易 1 截图", "验证交易 2 截图", "备注"]
    
    for col, header in enumerate(headers1, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    set_column_widths(ws, [15, 25, 18, 18, 20])
    
    yes_no_dv = DataValidation(type="list", formula1='"有，无"', allow_blank=True)
    ws.add_data_validation(yes_no_dv)
    
    for row in range(5, 8):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            cell.font = NORMAL_FONT
            if col in [3, 4]:
                yes_no_dv.add(cell)
    
    # === 公共接口修改 ===
    ws.merge_cells('A10:F10')
    section2 = ws.cell(row=10, column=1, value="【公共接口修改检查】")
    section2.fill = LIGHT_BLUE_FILL
    section2.font = Font(name="微软雅黑", size=11, bold=True)
    section2.alignment = LEFT_ALIGN
    section2.border = THIN_BORDER
    
    headers2 = ["接口名称", "影响交易列表", "单元测试（行覆盖）", "回归交易截图", "备注"]
    
    for col, header in enumerate(headers2, 1):
        cell = ws.cell(row=11, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    unit_dv = DataValidation(type="list", formula1='"已完成，未完成"', allow_blank=True)
    ws.add_data_validation(unit_dv)
    
    for row in range(12, 15):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            cell.font = NORMAL_FONT
            if col == 3:
                unit_dv.add(cell)
    
    # === QA 检查结果 ===
    ws.merge_cells('A17:F17')
    section3 = ws.cell(row=17, column=1, value="【冒烟准入 QA 检查结果】")
    section3.fill = LIGHT_BLUE_FILL
    section3.font = Font(name="微软雅黑", size=11, bold=True)
    section3.alignment = LEFT_ALIGN
    section3.border = THIN_BORDER
    
    headers3 = ["检查项目", "检查结果", "问题描述"]
    
    for col, header in enumerate(headers3, 1):
        cell = ws.cell(row=18, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    set_column_widths(ws, [35, 15, 40])
    
    check_items = [
        "版本单信息完整",
        "文件修改/新增清单完整",
        "接口清单完整准确",
        "接口改动已完成单元测试",
        "所有交易截图齐全",
        "报文/流水截图齐全",
        "公共组件/接口影响清单已梳理",
        "公共组件≥2 笔交易验证截图",
        "代码编译正常、无阻塞问题"
    ]
    
    result_dv = DataValidation(type="list", formula1='"通过，不通过"', allow_blank=True)
    ws.add_data_validation(result_dv)
    
    for row, item in enumerate(check_items, 19):
        cell = ws.cell(row=row, column=1, value=item)
        cell.border = THIN_BORDER
        cell.alignment = LEFT_ALIGN
        cell.font = NORMAL_FONT
        
        result_cell = ws.cell(row=row, column=2)
        result_cell.border = THIN_BORDER
        result_cell.alignment = CENTER_ALIGN
        result_cell.font = NORMAL_FONT
        result_dv.add(result_cell)
        
        desc_cell = ws.cell(row=row, column=3)
        desc_cell.border = THIN_BORDER
        desc_cell.alignment = LEFT_ALIGN
        desc_cell.font = NORMAL_FONT
    
    # 最终结论
    ws.merge_cells('A28:B28')
    conclusion = ws.cell(row=28, column=1, value="QA 最终结论：")
    conclusion.font = Font(name="微软雅黑", size=11, bold=True)
    conclusion.alignment = RIGHT_ALIGN
    
    conclusion_dv = DataValidation(type="list", formula1='"□ 通过 允许冒烟，□ 不通过 打回整改"', allow_blank=True)
    ws.add_data_validation(conclusion_dv)
    conclusion_cell = ws.cell(row=28, column=3)
    conclusion_cell.border = THIN_BORDER
    conclusion_cell.alignment = CENTER_ALIGN
    conclusion_dv.add(conclusion_cell)
    
    # 签字和日期
    ws.cell(row=30, column=1, value="QA 签字：").font = NORMAL_FONT
    ws.cell(row=30, column=3, value="检查日期：").font = NORMAL_FONT
    
    ws.freeze_panes = 'A4'


def main():
    """主函数"""
    wb = Workbook()
    
    # 创建 5 个工作表
    create_version_sheet(wb.active)
    create_files_sheet(wb.create_sheet())
    create_api_sheet(wb.create_sheet())
    create_test_sheet(wb.create_sheet())
    create_qa_sheet(wb.create_sheet())
    
    # 保存文件
    output_path = "/home/admin/.openclaw/workspace/lingxi-assistant/版本提测冒烟检查清单.xlsx"
    wb.save(output_path)
    print(f"✅ Excel 文件已生成：{output_path}")
    print("\n📋 包含 5 个工作表：")
    print("   1. 版本单")
    print("   2. 文件列表")
    print("   3. 接口清单")
    print("   4. 交易测试")
    print("   5. QA 检查")
    print("\n✨ 特性：")
    print("   - 蓝色主题美化")
    print("   - 自动换行")
    print("   - 下拉框选择（是/否、有/无、通过/不通过等）")
    print("   - 冻结首行")
    print("   - WPS 直接可用")


if __name__ == "__main__":
    main()
